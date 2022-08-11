import pandas as pd
from openpyxl import load_workbook

data = ["CONTROL #", "STUDENT TYPE", "LEVEL APPLIED FOR", 
"STRAND", "SCHOOL YEAR", "REGISTRATION DATE", "LEARNER REFERENCE NUMBER(LRN)",
"SURNAME", "FIRST NAME", "MIDDLE NAME", "BIRTHDATE", "CITIZENSHIP", "RELIGION",
 "PLACE OF BIRTH", "TELEPHONE NO.", "CELLPHONE NO.", "GENDER", "GOOGLE ACCOUNT",
 "HOME ADDRESS", "LAST SCHOOL ATTENDED", "GEN. AVERAGE", "ADDRESS OF LAST SCHOOL ATTENDED",
 "HONORS RECEIVED", "EDUCATION LEVEL", "PAYMENT TYPE", "OCCUPATION", "TOTAL FAMILY MONTHLY INCOME",
 "NUMBER OF SIBLINGS", "GUARDIAN'S NAME", "FAMILY STATUS", "DISCOUNT TYPE", "DOCUMENTS SUBMITTED",
 "Verified for completeness by Registrar:"]

workbook = load_workbook(filename="enrolled_student.xlsx")
lst = workbook.sheetnames
print(lst)
#df = pd.DataFrame(lst)
values = []
for sheet_name in lst:
    df = pd.read_excel("enrolled_student.xlsx", sheet_name=sheet_name, header=None, names=data)
    col_list = df["GENDER"].values.tolist()
    male = col_list.count("Male")
    female = col_list.count("Female")
    total = male + female
    values.append({'Section/Level':sheet_name, 'Male' : male, 'Female': female, 'Total' : total})

df = pd.DataFrame.from_dict(values)
df.sort_values('Section/Level')
df.to_excel("output.xlsx")  