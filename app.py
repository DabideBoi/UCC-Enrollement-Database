from flask import Flask, redirect, url_for, render_template, request
import pandas as pd
from openpyxl import load_workbook
import csv
import openpyxl
import os
import dashboard

"'temp-' + str(numz) +'.csv'" #temp file

data = ["CONTROL #", "STUDENT TYPE", "LEVEL APPLIED FOR", 
"STRAND", "SCHOOL YEAR", "REGISTRATION DATE", "LEARNER REFERENCE NUMBER(LRN)",
"SURNAME", "FIRST NAME", "MIDDLE NAME", "BIRTHDATE", "CITIZENSHIP", "RELIGION",
 "PLACE OF BIRTH", "TELEPHONE NO.", "CELLPHONE NO.", "GENDER", "GOOGLE ACCOUNT",
 "HOME ADDRESS", "LAST SCHOOL ATTENDED", "GEN. AVERAGE", "ADDRESS OF LAST SCHOOL ATTENDED",
 "HONORS RECEIVED", "EDUCATION LEVEL", "PAYMENT TYPE", "OCCUPATION", "TOTAL FAMILY MONTHLY INCOME",
 "NUMBER OF SIBLINGS", "GUARDIAN'S NAME", "FAMILY STATUS", "DISCOUNT TYPE", "DOCUMENTS SUBMITTED",
 "Verified for completeness by Registrar:"]


app = Flask(__name__)


def trans_gender(df, numz, value):
    df.loc[df['CONTROL #'] == numz, 'GENDER'] = value
    return df

@app.route("/")
def home():
    return render_template("base.html")

@app.route("/login", methods=["POST", "GET"])
def login():
    if request.method == "POST":
        num = request.form.get('numz', type=int)
        return redirect(url_for('tables', numz=num))
    else:
        return render_template("login.html")

@app.route("/success")
def success():
    return """<h1>Student is successfully added to the database</h1>
               <input type="submit" onClick="myFunction()" value="Return to Main Page"/> 
               <script>
               function myFunction() {
                   window.location.href="/login";  
                }
                </script>"""

@app.route("/tables/<numz>", methods=["GET", "POST"])
def tables(numz):
    temp_file = "temp-" + str(numz) +  ".xlsx"

    #This is the method when the user clicks a button
    if request.method == "POST":
        #This creates a local temp file for a more efficient runtime
        df = pd.read_excel('temp-' + str(numz) +'.xlsx', header=None, names=data)
        df['CONTROL #'] = df['CONTROL #'].astype(str)
        df_ = df.loc[df['CONTROL #'] == numz]
        grade_level = str(df_['LEVEL APPLIED FOR'].values[0])
        strand = str(df_['STRAND'].values[0])
        strand = strand.partition(' (')[0]
        df_ = df.loc[df['CONTROL #'] == numz]
        rows = df_.values.tolist()
        workbook = load_workbook(filename="enrolled_student.xlsx")
        selected = request.form.get("sectionz", type=str)
        sheet_name = ''

        #The code will run if the user decide to change the student's gender to Male
        if request.form.get('gender') == 'Male':
            #Change the gender and delete the existing file to create a new (changed) temp file that is ready to be enrolled
            df_ = trans_gender(df_, numz, 'Male')
            rows = df_.values.tolist()
            os.remove(temp_file)
            wb = openpyxl.Workbook()
            wb.save(filename=temp_file)
            workbook = load_workbook(filename=temp_file)
            sheet_name = "Sheet"
            sheet = workbook.create_sheet(sheet_name)
            print(sheet_name)
            sheet.title = (sheet_name)
            sheet = workbook[sheet_name]
            sheet.delete_rows(2, sheet.max_row-1)
            for row in rows:
                    sheet.append(row)
            workbook = workbook.save(filename=temp_file)

            #Will return to the tables page with a new assigned gender
            if grade_level == 'Grade 11' or grade_level == 'Grade 12':
                return render_template('test.html',
                                    number=str(numz),
                                    data = df_.to_html(),
                                    grade_level = df_['LEVEL APPLIED FOR'].values[0],
                                    shs_strand = strand,
                                    sections = ["Public_1", "Public_2", "Public_3", "Public_4", "Public_5", "Public_6", "Public_7", "Public_8", "Public_9", "Public_10", 
                                                "Private_1", "Private_2", "Private_3", "Private_4", "Private_5"],
                                    visible =  "visible")
            else: 
                return render_template('test.html',
                                    number=str(numz),
                                    data = df_.to_html(),
                                    grade_level = df_['LEVEL APPLIED FOR'].values[0],
                                    shs_strand = '',
                                    sections = [],
                                    visible = "hidden")
            
        #The code will run if the user decide to change the student's gender to Female
        elif request.form.get('gender') == 'Female':
            #Change the gender and delete the existing file to create a new (changed) temp file that is ready to be enrolled
            df_ = trans_gender(df_, numz, 'Female')
            rows = df_.values.tolist()
            os.remove(temp_file)
            wb = openpyxl.Workbook()
            wb.save(filename=temp_file)
            workbook = load_workbook(filename=temp_file)
            sheet_name = "Sheet"
            sheet = workbook.create_sheet(sheet_name)
            print(sheet_name)
            sheet.title = (sheet_name)
            sheet = workbook[sheet_name]
            sheet.delete_rows(2, sheet.max_row-1)
            for row in rows:
                    sheet.append(row)
            workbook = workbook.save(filename=temp_file)

            #Will return to the tables page with a new assigned gender
            if grade_level == 'Grade 11' or grade_level == 'Grade 12':
                return render_template('test.html',
                                    number=str(numz),
                                    data = df_.to_html(),
                                    grade_level = df_['LEVEL APPLIED FOR'].values[0],
                                    shs_strand = strand,
                                    sections = ["Public_1", "Public_2", "Public_3", "Public_4", "Public_5", "Public_6", "Public_7", "Public_8", "Public_9", "Public_10", 
                                                "Private_1", "Private_2", "Private_3", "Private_4", "Private_5"],
                                    visible =  "visible")
            else: 
                return render_template('test.html',
                                    number=str(numz),
                                    data = df_.to_html(),
                                    grade_level = df_['LEVEL APPLIED FOR'].values[0],
                                    shs_strand = '',
                                    sections = [],
                                    visible = "hidden")
            
        #If there are no changes in gender this block of code will run
        else:
            if grade_level == 'Grade 11' or grade_level == 'Grade 12':
                sheet_name = str(grade_level + "-" + selected + " " + strand)
            
            #Sheet Naming Format
            if sheet_name in workbook.sheetnames or grade_level in workbook.sheetnames:
                if grade_level == 'Grade 11' or grade_level == 'Grade 12':
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook[grade_level]
            else:
                if grade_level == 'Grade 11' or grade_level == 'Grade 12':
                    sheet = workbook.create_sheet(sheet_name)
                    print(sheet_name)
                    sheet.title = (sheet_name)
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(grade_level)
                    print(grade_level)
                    sheet.title = (grade_level)
                    sheet = workbook[grade_level]
                

            #Appends given row into the sheet
            for row in rows:
                sheet.append(row)

            workbook.save(filename="enrolled_student.xlsx")
            os.remove(temp_file)
            dashboard.Dasher()
            return redirect(url_for('success'))
    else:
        df = pd.read_excel('https://api.onedrive.com/v1.0/shares/u!aHR0cHM6Ly8xZHJ2Lm1zL3gvcyFBZ2FzRjdIdUtlRG9nY0JTcXVXdDRCYUZYSWdtdkE_ZT1OeWhHMzE/root/content')
        df['CONTROL #'] = df['CONTROL #'].astype(str)
        df_ = df.loc[df['CONTROL #'] == numz]
        rows = df_.values.tolist()

        wb = openpyxl.Workbook()
        wb.save(filename=temp_file)
        workbook = load_workbook(filename=temp_file)
        sheet_name = "Sheet"
        sheet = workbook.create_sheet(sheet_name)
        print(sheet_name)
        sheet.title = (sheet_name)
        sheet = workbook[sheet_name]
        for row in rows:
                sheet.append(row)
        workbook = workbook.save(filename=temp_file)
            
        grade_level = str(df_['LEVEL APPLIED FOR'].values[0])
        strand = str(df_['STRAND'].values[0])
        df_ = df.loc[df['CONTROL #'] == numz]
        if grade_level == 'Grade 11' or grade_level == 'Grade 12':
            return render_template('test.html',
                                number=str(numz),
                                data = df_.to_html(),
                                grade_level = df_['LEVEL APPLIED FOR'].values[0],
                                shs_strand = strand,
                                sections = ["Public_1", "Public_2", "Public_3", "Public_4", "Public_5", "Public_6", "Public_7", "Public_8", "Public_9", "Public_10", 
                                            "Private_1", "Private_2", "Private_3", "Private_4", "Private_5"],
                                visible =  "visible")
        else: 
            return render_template('test.html',
                                number=str(numz),
                                data = df_.to_html(),
                                grade_level = df_['LEVEL APPLIED FOR'].values[0],
                                shs_strand = '',
                                sections = [],
                                visible = "hidden")
        

if __name__ == "__main__":
    app.run(debug=True)